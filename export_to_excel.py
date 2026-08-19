"""
export_to_excel.py
===================
Converts the CSVs written by `master_pipeline.py --live` into ONE Excel
workbook, ready for modelling and with the requested charts built
in. Run this AFTER a --live run (it just reads the CSVs already in
outputs/, no new API calls):


    py -3.11 export_to_excel.py --ticker AAPL-US


Produces outputs/live_AAPL_US_model.xlsx with the raw, scoring, chart, rollup, trend, heatmap, and decision sheets:


  Raw Estimates -- every analyst estimate snapshot exported by the live pipeline,
                  including estimate value, revision date, snapshot date, broker,
                  actual EPS, price and market cap.
  Raw Actuals   -- one row per earnings quarter with actual EPS and report date.
  Raw Prices    -- one row per earnings quarter with point-in-time price, market
                  cap, shares outstanding and the exact price date used in Eq. 1.
  Raw Data      -- normalized (analyst, quarter) forecast-error rows used by the
                  existing modelling formulas.
  Leaderboard  -- one row per analyst. n_observations, accuracy_score,
                  consistency_score, avg_staleness_days/freshness_score (when
                  the pull captured analyst revision dates), their z-scores,
                  and the composite partial_reliability_score are all REAL
                  EXCEL FORMULAS referencing the Raw Data sheet (COUNTIFS/
                  SUMPRODUCT/AVERAGEIF/AVERAGE/STDEV) -- not pasted-in
                  numbers -- so the sheet recalculates if you edit/append Raw
                  Data. A bar chart of the top/bottom analysts sits next to
                  the table.
  Quarterly    -- the firm-level consensus forecast error per quarter
                  (from run_pipeline()'s consensus/industry output) with a
                  line chart showing the trend over time.
  Analyst Charts -- ONLY the analyst bubble map: accuracy vs. consistency,
                  bubble size = n_observations. No other analyst charts are added.
  Broker Rollup -- one row per brokerage, REAL FORMULAS (COUNTIF/AVERAGEIF)
                  averaging the Leaderboard's scores by current_broker, with
                  a bar chart -- which HOUSES field the more reliable
                  analysts on this name, not just which individuals.
  Analyst Trends -- forecast error over time for the top analysts (capped,
                  printed when it triggers, to keep the chart legible), REAL
                  FORMULAS (AVERAGEIFS) referencing Raw Data, one line series
                  per analyst.
  FE Heatmap   -- analyst x quarter grid (REAL FORMULAS, AVERAGEIFS) with a
                  red/white/blue conditional-formatting color scale -- shows
                  coverage gaps and over/under-estimation patterns across
                  every analyst at a glance.


This mirrors exactly what simple_accuracy_leaderboard() in
master_pipeline.py computes in Python -- see that function's docstring for
what accuracy/consistency do and don't capture (no NN/predictability
component yet; see fetch_live_ticker_data()'s KNOWN LIMITATION note for
why a single ticker can't support that yet).
"""


from __future__ import annotations


import argparse
import os


import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, BubbleChart, LineChart, Reference, ScatterChart, Series
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
BODY_FONT = Font(name="Arial")




def _style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL




def _autofit(ws, ncols, widths=None):
    for c in range(1, ncols + 1):
        letter = get_column_letter(c)
        ws.column_dimensions[letter].width = (widths or {}).get(c, 16)




def _apply_decision_theme(ws, title_range, header_row, ncols, widths):
    """Apply the visual language used by the supplied improved workbook."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f"A{header_row + 1}"
    ws.merge_cells(title_range)
    title = ws[title_range.split(":")[0]]
    title.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title.fill = HEADER_FILL
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24
    for c in range(1, ncols + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 34
    for c, width in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = width




def _style_decision_note(ws, row, text, end_col):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Arial", italic=True, size=10, color="404040")
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 30




def _add_scoring_weights(ws, start_col=18, start_row=4):
    """Create editable scoring weights without colliding with the main table."""
    labels = [("Accuracy", 0.50), ("Consistency", 0.25), ("Freshness", 0.15), ("Evidence", 0.10)]
    label_col = get_column_letter(start_col)
    value_col = get_column_letter(start_col + 1)
    ws.cell(row=start_row, column=start_col, value="Scoring weights")
    ws.cell(row=start_row, column=start_col).font = Font(name="Arial", bold=True, color="FFFFFF")
    ws.cell(row=start_row, column=start_col).fill = HEADER_FILL
    ws.cell(row=start_row, column=start_col + 1).fill = HEADER_FILL
    for i, (label, value) in enumerate(labels, start=1):
        r = start_row + i
        ws.cell(row=r, column=start_col, value=label).font = Font(name="Arial", color="404040")
        v = ws.cell(row=r, column=start_col + 1, value=value)
        v.font = Font(name="Arial", color="0000FF")
        v.number_format = '0%'
    ws.cell(row=start_row + 5, column=start_col, value="Total")
    ws.cell(row=start_row + 5, column=start_col).font = Font(name="Arial", bold=True, color="404040")
    total = ws.cell(row=start_row + 5, column=start_col + 1, value=f"=SUM({value_col}{start_row + 1}:{value_col}{start_row + 4})")
    total.font = Font(name="Arial", bold=True, color="000000")
    total.number_format = '0%'
    dv = DataValidation(type="decimal", operator="between", formula1="0", formula2="1", allow_blank=False)
    dv.promptTitle = "Scoring weight"
    dv.prompt = "Enter a decimal between 0 and 1. The four weights should total 100%."
    dv.error = "Weight must be between 0% and 100%."
    dv.errorTitle = "Invalid weight"
    ws.add_data_validation(dv)
    dv.add(f"{value_col}{start_row + 1}:{value_col}{start_row + 4}")
    ws.column_dimensions[label_col].width = 18
    ws.column_dimensions[value_col].width = 12
    return {
        "accuracy": f"${value_col}${start_row + 1}",
        "consistency": f"${value_col}${start_row + 2}",
        "freshness": f"${value_col}${start_row + 3}",
        "evidence": f"${value_col}${start_row + 4}",
        "total": f"${value_col}${start_row + 5}",
    }




def build_workbook(ticker: str) -> str:
    safe = ticker.replace("-", "_")
    outdir = "outputs"
    raw_path = f"{outdir}/live_{safe}_raw_forecast_errors.csv"
    raw_estimates_path = f"{outdir}/live_{safe}_raw_estimates.csv"
    raw_actuals_path = f"{outdir}/live_{safe}_raw_actuals.csv"
    raw_prices_path = f"{outdir}/live_{safe}_raw_prices.csv"
    consensus_path = f"{outdir}/live_{safe}_consensus.csv"
    run_info_path = f"{outdir}/live_{safe}_run_info.csv"


    if not os.path.exists(raw_path):
        raise FileNotFoundError(
            f"{raw_path} not found -- run `py -3.11 master_pipeline.py --live --ticker {ticker}` first."
        )


    raw = pd.read_csv(raw_path)
    raw_estimates = pd.read_csv(raw_estimates_path) if os.path.exists(raw_estimates_path) else pd.DataFrame()
    raw_actuals = pd.read_csv(raw_actuals_path) if os.path.exists(raw_actuals_path) else pd.DataFrame()
    raw_prices = pd.read_csv(raw_prices_path) if os.path.exists(raw_prices_path) else pd.DataFrame()
    consensus = pd.read_csv(consensus_path) if os.path.exists(consensus_path) else pd.DataFrame()
    run_info = pd.read_csv(run_info_path) if os.path.exists(run_info_path) else pd.DataFrame()


    # Order analysts by the SAME logic as simple_accuracy_leaderboard() in
    # master_pipeline.py, just to decide ROW ORDER -- the actual numbers in
    # the sheet are Excel formulas, computed fresh by Excel itself.
    import sys
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from master_pipeline import simple_accuracy_leaderboard
    from src.config import CREDIBILITY_PRIOR_OBS


    ordered = simple_accuracy_leaderboard(raw)
    analyst_order = ordered["analyst"].tolist()


    wb = Workbook()


    # ---------------------------------------------------------------
    # Sheet 0: Run Info
    # ---------------------------------------------------------------
    ws_info = wb.active
    ws_info.title = "Run Info"
    ws_info.sheet_view.showGridLines = False
    ws_info["A1"] = "Run Info"
    ws_info["A1"].font = Font(name="Arial", size=14, bold=True)
    if not run_info.empty:
        r = run_info.iloc[0]
        rows = [
            ("Ticker", r.get("ticker", ticker)),
            ("Quarters requested", r.get("quarters_requested", "")),
            ("Quarters actually pulled", r.get("quarters_pulled", "")),
            ("Earliest quarter in data", r.get("earliest_quarter", "")),
            ("Latest quarter in data", r.get("latest_quarter", "")),
            ("Number of analysts covered", r.get("n_analysts", "")),
            ("Total (analyst, quarter) observations", r.get("n_observations", "")),
            ("Pipeline run on", r.get("run_timestamp", "")),
        ]
    else:
        fallback_quarters = sorted(raw["quarter"].dropna().astype(str).unique().tolist()) if "quarter" in raw.columns else []
        rows = [
            ("Ticker", ticker),
            ("Quarters requested", ""),
            ("Quarters actually pulled", len(fallback_quarters)),
            ("Earliest quarter in data", fallback_quarters[0] if fallback_quarters else ""),
            ("Latest quarter in data", fallback_quarters[-1] if fallback_quarters else ""),
            ("Number of analysts covered", raw["analyst"].nunique() if "analyst" in raw.columns else ""),
            ("Total (analyst, quarter) observations", len(raw)),
            ("Pipeline run on", ""),
        ]
    for i, (label, value) in enumerate(rows, start=3):
        ws_info.cell(row=i, column=1, value=label).font = Font(name="Arial", bold=True)
        ws_info.cell(row=i, column=2, value=value)
    ws_info.column_dimensions["A"].width = 34
    ws_info.column_dimensions["B"].width = 30
    note_row = len(rows) + 5
    ws_info.cell(
        row=note_row, column=1,
        value="This workbook is formula driven. Edit Raw Data or the scoring weights and the decision sheets recalculate in Excel. The Raw Estimates, Raw Actuals and Raw Prices sheets are the audit layer from the live FactSet pull."
    ).font = Font(name="Arial", italic=True)


    # ---------------------------------------------------------------
    # Raw audit sheets: preserve the underlying data exported by the live
    # pipeline BEFORE the normalized modelling table. These are intentionally
    # pasted values, not formulas: they are the audit trail of the FactSet pull.
    # ---------------------------------------------------------------
    def _write_raw_input_sheet(name, df, title, widths=None):
        ws = wb.create_sheet(name)
        ws.sheet_view.showGridLines = False
        ws["A1"] = title
        ws["A1"].font = Font(name="Arial", size=14, bold=True)
        if df.empty:
            ws["A3"] = "No data file was produced for this run."
            ws["A3"].font = Font(name="Arial", italic=True, color="C00000")
            return ws
        df = df.copy()
        headers = list(df.columns)
        for c, h in enumerate(headers, start=1):
            ws.cell(row=3, column=c, value=h)
        _style_header(ws, len(headers), row=3)
        for r_idx, row_values in enumerate(df.itertuples(index=False, name=None), start=4):
            for c_idx, value in enumerate(row_values, start=1):
                ws.cell(row=r_idx, column=c_idx, value=None if pd.isna(value) else value)
        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{len(df) + 3}"
        _autofit(ws, len(headers), widths or {})
        return ws


    _write_raw_input_sheet(
        "Raw Estimates", raw_estimates,
        "Raw Estimates | analyst estimate snapshots and audit fields",
        {1: 14, 2: 14, 3: 10, 4: 8, 5: 12, 6: 12, 7: 22, 8: 24, 9: 14, 10: 16, 11: 14, 12: 16, 13: 18, 14: 16},
    )
    _write_raw_input_sheet(
        "Raw Actuals", raw_actuals,
        "Raw Actuals | reported earnings data",
        {1: 14, 2: 14, 3: 10, 4: 8, 5: 14, 6: 14, 7: 12},
    )
    _write_raw_input_sheet(
        "Raw Prices", raw_prices,
        "Raw Prices | point-in-time price and market data used in the model",
        {1: 14, 2: 14, 3: 10, 4: 8, 5: 14, 6: 14, 7: 16, 8: 16, 9: 18, 10: 20},
    )


    # ---------------------------------------------------------------
    # Sheet 1: Raw Data (the actual normalized FactSet modelling input)
    #
    # Sorted ascending by quarter (oldest first) -- this isn't cosmetic, the
    # Leaderboard's current_broker formula below relies on it: the
    # LOOKUP(2,1/(...)) trick returns the LAST matching row for an analyst,
    # which is only "her most recent broker" if rows are in chronological
    # order to begin with.
    # ---------------------------------------------------------------
    raw = raw.sort_values("quarter").reset_index(drop=True)
    ws_raw = wb.create_sheet("Raw Data")
    raw_cols = [
        "analyst", "broker", "broker_code", "firm", "quarter", "year", "fe", "market_cap", "sic_code",
        "staleness_days",
    ]
    raw_cols = [c for c in raw_cols if c in raw.columns]
    ws_raw.append(raw_cols)
    _style_header(ws_raw, len(raw_cols))
    for _, r in raw.iterrows():
        # Sanitize NaN -> None (a blank cell) rather than writing the literal
        # float nan -- matters most for staleness_days, which is legitimately
        # missing on some rows (analyst had no MODDATEN on file that
        # quarter), and AVERAGEIF/AVERAGE below rely on those being true
        # blanks, not the text "nan", to skip correctly.
        ws_raw.append([None if pd.isna(r[c]) else r[c] for c in raw_cols])
    _autofit(ws_raw, len(raw_cols), {1: 22, 2: 20, 3: 14, 4: 12, 5: 10, 6: 8, 7: 10, 8: 14, 9: 10, 10: 14})
    n_raw = len(raw) + 1  # +1 for header row
    has_broker = "broker" in raw_cols
    has_broker_code = "broker_code" in raw_cols
    has_staleness = "staleness_days" in raw_cols and raw["staleness_days"].notna().any()
    # de-dup basis for n_brokers: prefer the stable numeric code (matches
    # master_pipeline.py's _broker_fields(), which does the same) -- free-text
    # broker names can drift in spelling/punctuation across snapshots and
    # would over-count distinct brokers.
    dedup_field = "broker_code" if has_broker_code else ("broker" if has_broker else None)


    # ---------------------------------------------------------------
    # Sheet 2: Leaderboard (FORMULAS referencing Raw Data, not hardcoded)
    #
    # Column letters are computed from `headers` below (via lb_col/lb_letter)
    # rather than hand-counted -- a previous version of this file hardcoded
    # letters directly in the formula strings, and a stale-variable bug slipped
    # through a clean recalc.py run because of it (caught only by comparing
    # recalculated VALUES against known-correct Python output). Adding the
    # freshness columns here would have meant re-counting every letter after
    # column E by hand again, so this rewrite removes that whole class of bug.
    # ---------------------------------------------------------------
    ws_lb = wb.create_sheet("Leaderboard")
    headers = ["analyst", "rank", "n_observations", "mean_fe", "accuracy_score", "consistency_score"]
    if has_staleness:
        headers += ["avg_staleness_days", "freshness_score"]
    headers += ["accuracy_score_z", "consistency_score_z"]
    if has_staleness:
        headers.append("freshness_score_z")
    headers += ["partial_reliability_score_raw", "credibility_weight", "partial_reliability_score"]
    if has_broker:
        headers.append("current_broker")
    if has_broker_code:
        headers.append("current_broker_code")
    if dedup_field:
        headers.append("n_brokers")
    ws_lb.append(headers)
    _style_header(ws_lb, len(headers))


    # lb_col/lb_letter: Leaderboard sheet's OWN column index/letter for each
    # header name (NOT to be confused with `col`, which is Raw Data's).
    lb_col = {name: i + 1 for i, name in enumerate(headers)}
    lb_letter = {name: get_column_letter(i + 1) for i, name in enumerate(headers)}


    # Column letters in Raw Data (depends on which optional columns exist there).
    col = {name: chr(ord("A") + raw_cols.index(name)) for name in raw_cols}


    first_data_row = 2
    for i, analyst in enumerate(analyst_order):
        row = first_data_row + i
        a_col = f"$A${row}"
        # Raw Data ranges (fixed, absolute)
        rd_analyst = f"'Raw Data'!${col['analyst']}$2:${col['analyst']}${n_raw}"
        rd_fe = f"'Raw Data'!${col['fe']}$2:${col['fe']}${n_raw}"


        n_obs_ref = f"{lb_letter['n_observations']}{row}"
        mean_fe_ref = f"{lb_letter['mean_fe']}{row}"


        ws_lb.cell(row=row, column=lb_col["analyst"], value=analyst)
        ws_lb.cell(row=row, column=lb_col["n_observations"],
                   value=f"=COUNTIFS({rd_analyst},{a_col})")
        ws_lb.cell(row=row, column=lb_col["mean_fe"],
                   value=f"=AVERAGEIF({rd_analyst},{a_col},{rd_fe})")
        ws_lb.cell(
            row=row, column=lb_col["accuracy_score"],
            value=f"=-SUMPRODUCT(({rd_analyst}={a_col})*ABS({rd_fe}))/{n_obs_ref}",
        )  # accuracy_score = -mean(|fe|)
        ws_lb.cell(
            row=row, column=lb_col["consistency_score"],
            value=(
                f'=IFERROR(IF({n_obs_ref}>1,-SQRT(MAX(0,SUMPRODUCT(({rd_analyst}={a_col})'
                f'*(({rd_fe}-{mean_fe_ref})^2))/{n_obs_ref})),""),"")'
            ),
        )  # consistency_score = -stdev(fe), blank if only 1 observation or malformed row


        if has_staleness:
            rd_stale = f"'Raw Data'!${col['staleness_days']}$2:${col['staleness_days']}${n_raw}"
            avg_stale_ref = f"{lb_letter['avg_staleness_days']}{row}"
            ws_lb.cell(
                row=row, column=lb_col["avg_staleness_days"],
                value=f'=IFERROR(AVERAGEIF({rd_analyst},{a_col},{rd_stale}),"")',
            )  # AVERAGEIF ignores blank cells in the average range, so rows with
               # no revision_date on file that quarter don't skew this.
            ws_lb.cell(
                row=row, column=lb_col["freshness_score"],
                value=f'=IF(ISNUMBER({avg_stale_ref}),-{avg_stale_ref},"")',
            )  # higher = fresher = better, same sign convention as accuracy/consistency


    last_row = first_data_row + len(analyst_order) - 1


    def _z_formula(source_col_name: str) -> str:
        letter = lb_letter[source_col_name]
        rng = f"${letter}${first_data_row}:${letter}${last_row}"
        return (
            f'=IFERROR(IF(ISNUMBER({letter}{{row}}),({letter}{{row}}-AVERAGE({rng}))'
            f'/STDEV({rng}),""),"")'
        )


    z_pairs = [("accuracy_score", "accuracy_score_z"), ("consistency_score", "consistency_score_z")]
    if has_staleness:
        z_pairs.append(("freshness_score", "freshness_score_z"))
    z_score_col_names = [z_col for _, z_col in z_pairs]


    for row in range(first_data_row, last_row + 1):
        n_obs_ref = f"{lb_letter['n_observations']}{row}"  # fresh per row -- NOT the
                                                            # variable from the earlier,
                                                            # separate per-analyst loop
                                                            # above (that one's scoped to
                                                            # its own loop and unrelated;
                                                            # this is the exact class of
                                                            # stale-variable bug caught
                                                            # once before in this file).


        for source_col, z_col in z_pairs:
            ws_lb.cell(row=row, column=lb_col[z_col], value=_z_formula(source_col).format(row=row))


        # partial_reliability_score_raw: AVERAGE() over whichever z-scores are
        # actually numbers -- it silently skips blank/"" arguments on its own,
        # so this needs no special-casing for single-observation analysts
        # (blank consistency_score_z) or missing freshness data, mirroring
        # exactly how master_pipeline.py's out[z_cols].mean(axis=1) behaves in
        # Python (pandas skips NaN the same way).
        z_refs = ", ".join(f"{lb_letter[z_col]}{row}" for z_col in z_score_col_names)
        ws_lb.cell(
            row=row, column=lb_col["partial_reliability_score_raw"],
            value=f'=IFERROR(AVERAGE({z_refs}),"")',
        )
        ws_lb.cell(
            row=row, column=lb_col["credibility_weight"],
            value=f"={n_obs_ref}/({n_obs_ref}+{CREDIBILITY_PRIOR_OBS})",
        )  # credibility_weight -- n_observations / (n_observations + prior).
           # Same constant as master_pipeline.py's config.CREDIBILITY_PRIOR_OBS,
           # imported above rather than re-typed, so this can't silently drift
           # out of sync with the Python scoring logic it's meant to mirror.
        raw_ref = f"{lb_letter['partial_reliability_score_raw']}{row}"
        weight_ref = f"{lb_letter['credibility_weight']}{row}"
        ws_lb.cell(
            row=row, column=lb_col["partial_reliability_score"],
            value=f"={raw_ref}*{weight_ref}",
        )  # partial_reliability_score = raw score shrunk toward 0 by how little
           # history backs it -- THIS is what the sheet is meant to be read/sorted by.


        # rank: classic RANK() (NOT RANK.EQ -- that's a post-2007 function and
        # needs an "_xlfn." prefix openpyxl doesn't add automatically; without
        # it, LibreOffice's recalc engine throws #NAME? -- caught by running
        # recalc.py on this exact addition before shipping it) over the WHOLE
        # partial_reliability_score column (absolute range, every row) -- 1 =
        # best. Descending (order=0), since higher partial_reliability_score
        # is always better here. Ties get the same rank with the next rank
        # number skipped (RANK's standard "competition ranking" behavior,
        # e.g. 1, 2, 2, 4). This is purely a RE-EXPRESSION of
        # partial_reliability_score as a 1st/2nd/3rd... position -- it
        # doesn't change what's being measured, just how it's displayed, per
        # Patrick's request not to touch the underlying scoring.
        score_col_letter = lb_letter["partial_reliability_score"]
        score_range = f"${score_col_letter}${first_data_row}:${score_col_letter}${last_row}"
        score_ref = f"{score_col_letter}{row}"
        ws_lb.cell(
            row=row, column=lb_col["rank"],
            value=f"=RANK({score_ref},{score_range},0)",
        )


        if has_broker or has_broker_code or dedup_field:
            this_row_a_col = f"$A${row}"


            if has_broker:
                rd_broker = f"'Raw Data'!${col['broker']}$2:${col['broker']}${n_raw}"
                # current_broker: the classic LOOKUP(2,1/(condition),...) trick --
                # returns the LAST row in Raw Data matching this analyst, which is
                # her MOST RECENT broker because Raw Data is sorted ascending by
                # quarter above. Legacy-function-only (no dynamic arrays), no
                # CSE/array-entry required -- safe under LibreOffice's recalc.
                ws_lb.cell(
                    row=row, column=lb_col["current_broker"],
                    value=f'=IFERROR(LOOKUP(2,1/({rd_analyst}={this_row_a_col}),{rd_broker}),"")',
                )


            if has_broker_code:
                rd_code = f"'Raw Data'!${col['broker_code']}$2:${col['broker_code']}${n_raw}"
                ws_lb.cell(
                    row=row, column=lb_col["current_broker_code"],
                    value=f'=IFERROR(LOOKUP(2,1/({rd_analyst}={this_row_a_col}),{rd_code}),"")',
                )


            if dedup_field:
                # n_brokers: standard SUMPRODUCT "count distinct values matching a
                # criteria" formula, run against whichever field is the reliable
                # de-dup key (broker_code if we have it, else broker name -- same
                # preference master_pipeline.py's _broker_fields() uses). For each
                # of this analyst's rows, add 1/(how many of her rows share that
                # SAME value); rows at a broker for k quarters contribute
                # k * (1/k) = 1, so the total is her count of DISTINCT brokers.
                rd_dedup = f"'Raw Data'!${col[dedup_field]}$2:${col[dedup_field]}${n_raw}"
                ws_lb.cell(
                    row=row, column=lb_col["n_brokers"],
                    value=(
                        f"=IF({n_obs_ref}=0,0,SUMPRODUCT(({rd_analyst}={this_row_a_col})"
                        f"/COUNTIFS({rd_analyst},{rd_analyst},{rd_dedup},{rd_dedup})))"
                    ),
                )


    _style_header(ws_lb, len(headers))
    widths = {lb_col["analyst"]: 22, lb_col["rank"]: 8, lb_col["n_observations"]: 14, lb_col["mean_fe"]: 10,
              lb_col["accuracy_score"]: 12, lb_col["consistency_score"]: 14}
    if has_staleness:
        widths[lb_col["avg_staleness_days"]] = 16
        widths[lb_col["freshness_score"]] = 14
    widths[lb_col["accuracy_score_z"]] = 12
    widths[lb_col["consistency_score_z"]] = 14
    if has_staleness:
        widths[lb_col["freshness_score_z"]] = 14
    widths[lb_col["partial_reliability_score_raw"]] = 22
    widths[lb_col["credibility_weight"]] = 16
    widths[lb_col["partial_reliability_score"]] = 20
    if has_broker:
        widths[lb_col["current_broker"]] = 20
    if has_broker_code:
        widths[lb_col["current_broker_code"]] = 16
    if dedup_field:
        widths[lb_col["n_brokers"]] = 10
    _autofit(ws_lb, len(headers), widths)


    # Bar chart: partial_reliability_score (credibility-weighted) per analyst
    chart = BarChart()
    chart.type = "col"
    chart.title = f"{ticker} -- analyst partial reliability score (credibility-weighted)"
    chart.y_axis.title = "z-scored composite x credibility weight"
    chart.x_axis.title = "analyst"
    data = Reference(ws_lb, min_col=lb_col["partial_reliability_score"], min_row=1, max_row=last_row)
    cats = Reference(ws_lb, min_col=1, min_row=first_data_row, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width, chart.height = 32, 14
    chart_anchor_col = get_column_letter(len(headers) + 2)  # 2-column gap past the last data column
    ws_lb.add_chart(chart, f"{chart_anchor_col}{first_data_row}")


    # ---------------------------------------------------------------
    # Sheet 3: Quarterly (firm-level consensus forecast error trend)
    # ---------------------------------------------------------------
    if not consensus.empty:
        ws_q = wb.create_sheet("Quarterly")
        q_cols = ["quarter", "consensus_fe", "market_cap", "industry"]
        ws_q.append(q_cols)
        _style_header(ws_q, len(q_cols))
        consensus_sorted = consensus.sort_values("quarter")
        for _, r in consensus_sorted.iterrows():
            ws_q.append([r["quarter"], r["consensus_fe"], r["market_cap"], r.get("industry", "")])
        _autofit(ws_q, len(q_cols), {1: 12, 2: 16, 3: 14, 4: 14})
        n_q = len(consensus_sorted) + 1


        line = LineChart()
        line.title = f"{ticker} -- consensus forecast error by quarter"
        line.y_axis.title = "consensus FE (fraction of price)"
        line.x_axis.title = "quarter"
        data = Reference(ws_q, min_col=2, min_row=1, max_row=n_q)
        cats = Reference(ws_q, min_col=1, min_row=2, max_row=n_q)
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        line.width, line.height = 24, 12
        ws_q.add_chart(line, "F2")


    # ---------------------------------------------------------------
    # Sheet 4: Analyst Charts -- keep ONLY the analyst bubble map.
    # This is the clearest view of accuracy vs. consistency, while bubble
    # size shows the amount of evidence behind each analyst.
    # ---------------------------------------------------------------
    ws_charts = wb.create_sheet("Analyst Charts")
    ws_charts["A1"] = f"{ticker} -- analyst bubble map (accuracy vs. consistency; bubble size = observations)"
    ws_charts["A1"].font = Font(name="Arial", bold=True, size=12)


    bubble = BubbleChart()
    bx = Reference(ws_lb, min_col=lb_col["accuracy_score"], min_row=first_data_row, max_row=last_row)
    by = Reference(ws_lb, min_col=lb_col["consistency_score"], min_row=first_data_row, max_row=last_row)
    bsize = Reference(ws_lb, min_col=lb_col["n_observations"], min_row=first_data_row, max_row=last_row)
    bubble.series.append(Series(by, xvalues=bx, zvalues=bsize, title="Analysts (bubble size = observations)"))
    bubble.title = f"{ticker} -- accuracy vs. consistency (bubble size = observations)"
    bubble.x_axis.title = "Accuracy score (higher = smaller average error)"
    bubble.y_axis.title = "Consistency score (higher = steadier)"
    bubble.width, bubble.height = 24, 14
    ws_charts.add_chart(bubble, "A3")


    # ---------------------------------------------------------------
    # Sheet 5: Broker Rollup -- REAL FORMULAS (COUNTIF/AVERAGEIF) averaging
    # the Leaderboard's own scores by brokerage -- which HOUSES field more
    # reliable analysts on this name, not just which individuals.
    # KNOWN EDGE CASE: COUNTIF/AVERAGEIF treat '*' and '?' in the criteria as
    # wildcards -- a broker name containing either literally would misfire.
    # Not handled here (Excel has no clean escape for this without a helper
    # column); flagging it rather than hiding it.
    # ---------------------------------------------------------------
    if has_broker:
        ws_broker = wb.create_sheet("Broker Rollup")
        broker_headers = ["broker", "n_analysts", "avg_accuracy_score", "avg_consistency_score",
                           "avg_partial_reliability_score"]
        ws_broker.append(broker_headers)
        _style_header(ws_broker, len(broker_headers))


        # Row order only (not the cell VALUES, which are live formulas below)
        # comes from the same Python leaderboard (`ordered`) used to order
        # analysts above -- ranks brokers by their analysts' average score.
        broker_avg = (
            ordered.dropna(subset=["current_broker"])
            .groupby("current_broker")["partial_reliability_score"]
            .mean()
            .sort_values(ascending=False)
        )
        broker_order = broker_avg.index.tolist()


        lb_broker_rng = f"Leaderboard!${lb_letter['current_broker']}${first_data_row}:${lb_letter['current_broker']}${last_row}"
        lb_acc_rng = f"Leaderboard!${lb_letter['accuracy_score']}${first_data_row}:${lb_letter['accuracy_score']}${last_row}"
        lb_cons_rng = f"Leaderboard!${lb_letter['consistency_score']}${first_data_row}:${lb_letter['consistency_score']}${last_row}"
        lb_score_rng = f"Leaderboard!${lb_letter['partial_reliability_score']}${first_data_row}:${lb_letter['partial_reliability_score']}${last_row}"


        first_broker_row = 2
        for i, broker in enumerate(broker_order):
            row = first_broker_row + i
            b_ref = f"$A${row}"
            ws_broker.cell(row=row, column=1, value=broker)
            ws_broker.cell(row=row, column=2, value=f"=COUNTIF({lb_broker_rng},{b_ref})")
            ws_broker.cell(row=row, column=3, value=f'=IFERROR(AVERAGEIF({lb_broker_rng},{b_ref},{lb_acc_rng}),"")')
            ws_broker.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIF({lb_broker_rng},{b_ref},{lb_cons_rng}),"")')
            ws_broker.cell(row=row, column=5, value=f'=IFERROR(AVERAGEIF({lb_broker_rng},{b_ref},{lb_score_rng}),"")')
        last_broker_row = first_broker_row + len(broker_order) - 1
        _autofit(ws_broker, len(broker_headers), {1: 24, 2: 12, 3: 18, 4: 20, 5: 26})


        broker_chart = BarChart()
        broker_chart.type = "col"
        broker_chart.title = f"{ticker} -- avg analyst reliability score by broker"
        broker_chart.y_axis.title = "avg partial_reliability_score"
        broker_chart.x_axis.title = "broker"
        bdata = Reference(ws_broker, min_col=5, min_row=1, max_row=last_broker_row)
        bcats = Reference(ws_broker, min_col=1, min_row=first_broker_row, max_row=last_broker_row)
        broker_chart.add_data(bdata, titles_from_data=True)
        broker_chart.set_categories(bcats)
        broker_chart.width, broker_chart.height = 28, 14
        ws_broker.add_chart(broker_chart, f"G{first_broker_row}")


    # ---------------------------------------------------------------
    # Sheet 6: Analyst Trends -- forecast error over time, top analysts by
    # score (capped so the chart stays legible -- the cap is PRINTED, never
    # silent). REAL FORMULAS (AVERAGEIFS) referencing Raw Data.
    # ---------------------------------------------------------------
    TREND_TOP_N = 8
    trend_analysts = analyst_order[:TREND_TOP_N]
    if len(analyst_order) > TREND_TOP_N:
        print(f"[export_to_excel] Analyst Trends: showing top {TREND_TOP_N} of "
              f"{len(analyst_order)} analysts (by partial_reliability_score) to keep the chart readable.")


    quarters_sorted = sorted(raw["quarter"].unique())
    rd_analyst2 = f"'Raw Data'!${col['analyst']}$2:${col['analyst']}${n_raw}"
    rd_quarter2 = f"'Raw Data'!${col['quarter']}$2:${col['quarter']}${n_raw}"
    rd_fe2 = f"'Raw Data'!${col['fe']}$2:${col['fe']}${n_raw}"


    ws_trend = wb.create_sheet("Analyst Trends")
    trend_headers = ["quarter"] + trend_analysts
    ws_trend.append(trend_headers)
    _style_header(ws_trend, len(trend_headers))
    for i, q in enumerate(quarters_sorted):
        row = 2 + i
        ws_trend.cell(row=row, column=1, value=q)
        q_ref = f"$A${row}"
        for j in range(len(trend_analysts)):
            header_ref = f"{get_column_letter(2 + j)}$1"
            ws_trend.cell(
                row=row, column=2 + j,
                value=f'=IFERROR(AVERAGEIFS({rd_fe2},{rd_analyst2},{header_ref},{rd_quarter2},{q_ref}),"")',
            )
    n_trend_rows = len(quarters_sorted) + 1
    _autofit(ws_trend, len(trend_headers), {1: 12})


    trend_chart = LineChart()
    trend_chart.title = f"{ticker} -- forecast error over time (top {len(trend_analysts)} analysts)"
    trend_chart.y_axis.title = "forecast error (fraction of price)"
    trend_chart.x_axis.title = "quarter"
    tdata = Reference(ws_trend, min_col=2, max_col=len(trend_headers), min_row=1, max_row=n_trend_rows)
    tcats = Reference(ws_trend, min_col=1, min_row=2, max_row=n_trend_rows)
    trend_chart.add_data(tdata, titles_from_data=True)
    trend_chart.set_categories(tcats)
    trend_chart.width, trend_chart.height = 32, 16
    ws_trend.add_chart(trend_chart, f"{get_column_letter(len(trend_headers) + 2)}1")


    # ---------------------------------------------------------------
    # Sheet 7: FE Heatmap -- analyst x quarter grid (REAL FORMULAS,
    # AVERAGEIFS) + a red/white/blue conditional-formatting color scale.
    # Shows coverage gaps (blank cells) and over/under-estimation patterns
    # across EVERY analyst at once, not just the top few.
    # ---------------------------------------------------------------
    ws_heat = wb.create_sheet("FE Heatmap")
    heat_headers = ["analyst"] + quarters_sorted
    ws_heat.append(heat_headers)
    _style_header(ws_heat, len(heat_headers))
    for i, analyst in enumerate(analyst_order):
        row = 2 + i
        ws_heat.cell(row=row, column=1, value=analyst)
        a_ref = f"$A${row}"
        for j, q in enumerate(quarters_sorted):
            q_header_ref = f"{get_column_letter(2 + j)}$1"
            ws_heat.cell(
                row=row, column=2 + j,
                value=f'=IFERROR(AVERAGEIFS({rd_fe2},{rd_analyst2},{a_ref},{rd_quarter2},{q_header_ref}),"")',
            )
    n_heat_rows = len(analyst_order) + 1
    _autofit(ws_heat, len(heat_headers), {1: 22})


    heat_range = f"B2:{get_column_letter(len(heat_headers))}{n_heat_rows}"
    color_rule = ColorScaleRule(
        start_type="min", start_color="F8696B",           # red = most negative (too pessimistic)
        mid_type="num", mid_value=0, mid_color="FFFFFF",  # white = exactly zero bias
        end_type="max", end_color="5A8AC6",                # blue = most positive (too optimistic)
    )
    ws_heat.conditional_formatting.add(heat_range, color_rule)


    # ---------------------------------------------------------------
    # Sheet 8: Analyst Decision -- decision-oriented view built from
    # Leaderboard. This is a presentation layer; it does not alter the
    # underlying Python scoring methodology.
    # ---------------------------------------------------------------
    ws_ad = wb.create_sheet("Analyst Decision")
    ws_ad.sheet_view.showGridLines = False
    ws_ad.merge_cells("A1:P1")
    ws_ad["A1"] = "Analyst Decision"
    ws_ad["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    ws_ad["A1"].fill = HEADER_FILL
    ws_ad["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws_ad.row_dimensions[1].height = 24
    ws_ad.merge_cells("A2:P2")
    ws_ad["A2"] = ("Purpose: identify the analysts whose historical forecast errors are most reliable, while explicitly "
                   "adjusting for consistency, freshness, and amount of evidence. Lower MAE and lower error volatility "
                   "are better. The final score is a weighted 0 to 100 score, then confidence-adjusted for sample size.")
    ws_ad["A2"].font = Font(name="Arial", size=10, color="404040")
    ws_ad["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_ad.row_dimensions[2].height = 36
    weights = _add_scoring_weights(ws_ad, start_col=18, start_row=4)
    ad_headers = [
        "Analyst", "Current broker", "Observations", "Avg absolute FE (MAE)", "Error volatility (stdev)",
        "Avg staleness (days)", "Accuracy score", "Consistency score", "Freshness score", "Evidence score",
        "Base decision score", "Confidence factor", "Adjusted decision score", "Rank", "Evidence flag", "Interpretation",
    ]
    ad_header_row = 6
    for c, h in enumerate(ad_headers, start=1):
        ws_ad.cell(row=ad_header_row, column=c, value=h)
    for c in range(1, len(ad_headers) + 1):
        cell = ws_ad.cell(row=ad_header_row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_ad.freeze_panes = "A7"
    ws_ad.auto_filter.ref = f"A6:P{6 + len(analyst_order)}"
    widths = [22, 28, 12, 18, 20, 18, 14, 16, 14, 13, 17, 15, 20, 8, 18, 28]
    for idx, width in enumerate(widths, start=1):
        ws_ad.column_dimensions[get_column_letter(idx)].width = width
    ws_ad.column_dimensions["R"].width = 18
    ws_ad.column_dimensions["S"].width = 12


    rd_analyst = f"'Raw Data'!${col['analyst']}$2:${col['analyst']}${n_raw}"
    rd_fe = f"'Raw Data'!${col['fe']}$2:${col['fe']}${n_raw}"
    rd_broker = f"'Raw Data'!${col['broker']}$2:${col['broker']}${n_raw}" if has_broker else None
    rd_stale = f"'Raw Data'!${col['staleness_days']}$2:${col['staleness_days']}${n_raw}" if has_staleness else None
    ad_first = 7
    ad_last = ad_first + len(analyst_order) - 1
    for i, analyst in enumerate(analyst_order):
        r = ad_first + i
        ws_ad.cell(r, 1, f"=Leaderboard!{lb_letter['analyst']}{first_data_row + i}")
        if has_broker:
            ws_ad.cell(r, 2, f"=Leaderboard!{lb_letter['current_broker']}{first_data_row + i}")
        ws_ad.cell(r, 3, f"=Leaderboard!{lb_letter['n_observations']}{first_data_row + i}")
        ws_ad.cell(r, 4, f"=-Leaderboard!{lb_letter['accuracy_score']}{first_data_row + i}")
        ws_ad.cell(r, 5, f'=IFERROR(-Leaderboard!{lb_letter["consistency_score"]}{first_data_row + i},"")')
        if has_staleness:
            ws_ad.cell(r, 6, f"=Leaderboard!{lb_letter['avg_staleness_days']}{first_data_row + i}")
        else:
            ws_ad.cell(r, 6, "")
        # Normalized 0-100 component scores. Blank raw metrics stay blank.
        ws_ad.cell(r, 7, f'=IF(ISNUMBER(D{r}),IFERROR(100*(MAX($D${ad_first}:$D${ad_last})-D{r})/(MAX($D${ad_first}:$D${ad_last})-MIN($D${ad_first}:$D${ad_last})),100),"")')
        ws_ad.cell(r, 8, f'=IF(ISNUMBER(E{r}),IFERROR(100*(MAX($E${ad_first}:$E${ad_last})-E{r})/(MAX($E${ad_first}:$E${ad_last})-MIN($E${ad_first}:$E${ad_last})),100),"")')
        if has_staleness:
            ws_ad.cell(r, 9, f'=IF(ISNUMBER(F{r}),IFERROR(100*(MAX($F${ad_first}:$F${ad_last})-F{r})/(MAX($F${ad_first}:$F${ad_last})-MIN($F${ad_first}:$F${ad_last})),100),"")')
        else:
            ws_ad.cell(r, 9, "")
        ws_ad.cell(r, 10, f'=IFERROR(MIN(100,C{r}/\'Run Info\'!$B$5*100),"")')
        # Use only the components that exist; missing freshness falls back to 3-part average plus evidence.
        if has_staleness:
            ws_ad.cell(r, 11, f'=IFERROR((G{r}*{weights["accuracy"]}+H{r}*{weights["consistency"]}+I{r}*{weights["freshness"]}+J{r}*{weights["evidence"]})/{weights["total"]},"")')
        else:
            ws_ad.cell(r, 11, f'=IFERROR((G{r}*{weights["accuracy"]}+H{r}*{weights["consistency"]}+J{r}*{weights["evidence"]})/({weights["accuracy"]}+{weights["consistency"]}+{weights["evidence"]}),"")')
        ws_ad.cell(r, 12, f'=IFERROR(C{r}/(C{r}+4),0)')
        ws_ad.cell(r, 13, f'=IFERROR(K{r}*L{r},"")')
        ws_ad.cell(r, 14, f'=IFERROR(RANK(M{r},$M${ad_first}:$M${ad_last},0),"")')
        ws_ad.cell(r, 15, f'=IF(C{r}<4,"Limited history",IF(C{r}<6,"Moderate history","Good history"))')
        ws_ad.cell(r, 16, f'=IF(C{r}<4,"Use cautiously: small sample",IF(M{r}>=70,"High confidence",IF(M{r}>=50,"Usable with context","Lower confidence")))')
        for c in range(1, 17):
            cell = ws_ad.cell(r, c)
            cell.font = Font(name="Arial", color="000000", bold=(c == 1))
            cell.alignment = Alignment(vertical="center", wrap_text=(c in (2, 15, 16)))
    for r in range(ad_first, ad_last + 1):
        ws_ad.cell(r, 3).number_format = '0'
        for c in (4, 5):
            ws_ad.cell(r, c).number_format = '0.0000'
        for c in (7, 8, 9, 10, 11, 13):
            ws_ad.cell(r, c).number_format = '0.0'
        ws_ad.cell(r, 12).number_format = '0.0%'
    _style_decision_note(ws_ad, ad_last + 2, "How to read this sheet", 16)
    notes = [
        "1. Accuracy uses mean absolute forecast error, so systematic bias does not cancel out.",
        "2. Consistency uses forecast error volatility. A lower standard deviation is better.",
        "3. Freshness rewards analysts whose observations are more recent.",
        "4. Evidence is capped at the number of quarters actually pulled, so thin coverage is penalized.",
        "5. The confidence factor deliberately shrinks small samples so a short lucky history cannot dominate a longer track record.",
    ]
    for j, text in enumerate(notes, start=ad_last + 3):
        _style_decision_note(ws_ad, j, text, 16)
    # Conditional formatting for decision score and evidence flag.
    ws_ad.conditional_formatting.add(f"M{ad_first}:M{ad_last}", ColorScaleRule(start_type="min", start_color="F8696B", mid_type="num", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"))
    ws_ad.conditional_formatting.add(f"O{ad_first}:O{ad_last}", CellIsRule(operator="equal", formula=['"Limited history"'], fill=PatternFill("solid", fgColor="FCE4D6")))


    # ---------------------------------------------------------------
    # Sheet 9: Broker Decision -- direct aggregation from Raw Data, not
    # a simple average of analyst scores. This keeps house-level results
    # meaningful when multiple analysts belong to the same broker.
    # ---------------------------------------------------------------
    if has_broker:
        ws_bd = wb.create_sheet("Broker Decision")
        ws_bd.sheet_view.showGridLines = False
        ws_bd.merge_cells("A1:O1")
        ws_bd["A1"] = "Broker Decision"
        ws_bd["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        ws_bd["A1"].fill = HEADER_FILL
        ws_bd["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws_bd.row_dimensions[1].height = 24
        ws_bd.merge_cells("A2:O2")
        ws_bd["A2"] = ("Purpose: evaluate brokers as a house-level forecasting source by aggregating all available observations directly "
                        "from Raw Data. This avoids simply averaging analyst scores and remains meaningful if a broker has multiple analysts.")
        ws_bd["A2"].font = Font(name="Arial", size=10, color="404040")
        ws_bd["A2"].alignment = Alignment(wrap_text=True, vertical="top")
        ws_bd.row_dimensions[2].height = 36
        bd_weights = _add_scoring_weights(ws_bd, start_col=17, start_row=4)
        bd_headers = [
            "Broker", "Analysts covered", "Observations", "Avg absolute FE (MAE)", "Error volatility (stdev)",
            "Avg staleness (days)", "Accuracy score", "Consistency score", "Freshness score", "Evidence score",
            "Base decision score", "Confidence factor", "Adjusted decision score", "Rank", "Interpretation",
        ]
        bd_header_row = 6
        for c, h in enumerate(bd_headers, start=1):
            ws_bd.cell(row=bd_header_row, column=c, value=h)
        for c in range(1, len(bd_headers) + 1):
            cell = ws_bd.cell(row=bd_header_row, column=c)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_bd.freeze_panes = "A7"
        ws_bd.auto_filter.ref = f"A6:O{6 + len(broker_order)}"
        bd_widths = [24, 16, 12, 18, 20, 18, 14, 16, 14, 13, 17, 15, 20, 8, 28]
        for idx, width in enumerate(bd_widths, start=1):
            ws_bd.column_dimensions[get_column_letter(idx)].width = width
        first_bd = 7
        last_bd = first_bd + len(broker_order) - 1
        rd_broker = f"'Raw Data'!${col['broker']}$2:${col['broker']}${n_raw}"
        for i, broker in enumerate(broker_order):
            r = first_bd + i
            ws_bd.cell(r, 1, f"='Broker Rollup'!A{2 + i}")
            ws_bd.cell(r, 2, f'=IFERROR(SUMPRODUCT(({rd_broker}=$A{r})/COUNTIFS({rd_broker},{rd_broker},{rd_analyst},{rd_analyst})),"")')
            ws_bd.cell(r, 3, f'=COUNTIF({rd_broker},$A{r})')
            ws_bd.cell(r, 4, f'=IFERROR(SUMPRODUCT(({rd_broker}=$A{r})*ABS({rd_fe}))/C{r},"")')
            ws_bd.cell(r, 5, f'=IF(C{r}>1,SQRT(SUMPRODUCT(({rd_broker}=$A{r})*({rd_fe}-AVERAGEIF({rd_broker},$A{r},{rd_fe}))^2)/C{r}),"")')
            if has_staleness:
                ws_bd.cell(r, 6, f'=IFERROR(AVERAGEIF({rd_broker},$A{r},{rd_stale}),"")')
            else:
                ws_bd.cell(r, 6, "")
            ws_bd.cell(r, 7, f'=IF(ISNUMBER(D{r}),IFERROR(100*(MAX($D${first_bd}:$D${last_bd})-D{r})/(MAX($D${first_bd}:$D${last_bd})-MIN($D${first_bd}:$D${last_bd})),100),"")')
            ws_bd.cell(r, 8, f'=IF(ISNUMBER(E{r}),IFERROR(100*(MAX($E${first_bd}:$E${last_bd})-E{r})/(MAX($E${first_bd}:$E${last_bd})-MIN($E${first_bd}:$E${last_bd})),100),"")')
            if has_staleness:
                ws_bd.cell(r, 9, f'=IF(ISNUMBER(F{r}),IFERROR(100*(MAX($F${first_bd}:$F${last_bd})-F{r})/(MAX($F${first_bd}:$F${last_bd})-MIN($F${first_bd}:$F${last_bd})),100),"")')
            else:
                ws_bd.cell(r, 9, "")
            ws_bd.cell(r, 10, f'=IFERROR(MIN(100,C{r}/\'Run Info\'!$B$5*100),"")')
            if has_staleness:
                ws_bd.cell(r, 11, f'=IFERROR((G{r}*{bd_weights["accuracy"]}+H{r}*{bd_weights["consistency"]}+I{r}*{bd_weights["freshness"]}+J{r}*{bd_weights["evidence"]})/{bd_weights["total"]},"")')
            else:
                ws_bd.cell(r, 11, f'=IFERROR((G{r}*{bd_weights["accuracy"]}+H{r}*{bd_weights["consistency"]}+J{r}*{bd_weights["evidence"]})/({bd_weights["accuracy"]}+{bd_weights["consistency"]}+{bd_weights["evidence"]}),"")')
            ws_bd.cell(r, 12, f'=IFERROR(C{r}/(C{r}+4),0)')
            ws_bd.cell(r, 13, f'=IFERROR(K{r}*L{r},"")')
            ws_bd.cell(r, 14, f'=IFERROR(RANK(M{r},$M${first_bd}:$M${last_bd},0),"")')
            ws_bd.cell(r, 15, f'=IF(C{r}<4,"Limited evidence",IF(M{r}>=70,"High confidence",IF(M{r}>=50,"Usable with context","Lower confidence")))')
            for c in range(1, 16):
                cell = ws_bd.cell(r, c)
                cell.font = Font(name="Arial", color="000000")
                cell.alignment = Alignment(vertical="center", wrap_text=(c == 15))
        for r in range(first_bd, last_bd + 1):
            for c in (4, 5):
                ws_bd.cell(r, c).number_format = '0.0000'
            for c in (7, 8, 9, 10, 11, 13):
                ws_bd.cell(r, c).number_format = '0.0'
            ws_bd.cell(r, 12).number_format = '0.0%'
        _style_decision_note(ws_bd, last_bd + 2, "Important interpretation note", 15)
        _style_decision_note(ws_bd, last_bd + 3, "Broker metrics are computed directly from Raw Data. When a broker has multiple analysts, this sheet aggregates the house's observations rather than averaging analyst scores.", 15)
        ws_bd.conditional_formatting.add(f"M{first_bd}:M{last_bd}", ColorScaleRule(start_type="min", start_color="F8696B", mid_type="num", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"))


    # ---------------------------------------------------------------
    # FF5 strategy outputs -- populated when master_pipeline.py was run
    # with --factor-model FF5 --factor-backtest. These are separate from
    # the analyst bubble map and do not add any chart to Analyst Charts.
    # ---------------------------------------------------------------
    ff_factors_path = f"{outdir}/live_{safe}_ff_factors.csv"
    strategy_returns_path = f"{outdir}/live_{safe}_strategy_returns.csv"
    factor_regression_path = f"{outdir}/live_{safe}_factor_regression.csv"


    ff_factors = pd.read_csv(ff_factors_path) if os.path.exists(ff_factors_path) else pd.DataFrame()
    strategy_returns = pd.read_csv(strategy_returns_path) if os.path.exists(strategy_returns_path) else pd.DataFrame()
    factor_regression = pd.read_csv(factor_regression_path) if os.path.exists(factor_regression_path) else pd.DataFrame()


    if not ff_factors.empty or not strategy_returns.empty or not factor_regression.empty:
        # FF5 Factors: show the pure five-factor panel only, even if an older
        # or alternate run left MOM in the exported factor CSV.
        ws_ff = wb.create_sheet("FF5 Factors")
        ws_ff.sheet_view.showGridLines = False
        ws_ff["A1"] = f"{ticker} -- Fama-French 5 Factors"
        ws_ff["A1"].font = Font(name="Arial", size=14, bold=True)
        ws_ff["A2"] = "Source: Ken French Fama-French 5-factor data already used by src/factors.py"
        ws_ff["A2"].font = Font(name="Arial", size=9, color="666666")
        required_factor_cols = ["date", "Mkt-RF", "SMB", "HML", "RMW", "CMA", "RF"]
        ff_cols = [c for c in required_factor_cols if c in ff_factors.columns]
        if ff_cols:
            ws_ff.append([])
            ws_ff.append(ff_cols)
            _style_header(ws_ff, len(ff_cols), row=4)
            for _, row in ff_factors[ff_cols].iterrows():
                ws_ff.append([row[c] for c in ff_cols])
            for col_idx in range(1, len(ff_cols) + 1):
                ws_ff.cell(4, col_idx).alignment = Alignment(horizontal="center")
            for r in range(5, 5 + len(ff_factors)):
                if "date" in ff_cols:
                    ws_ff.cell(r, ff_cols.index("date") + 1).number_format = "yyyy-mm-dd"
                for name in ["Mkt-RF", "SMB", "HML", "RMW", "CMA", "RF"]:
                    if name in ff_cols:
                        ws_ff.cell(r, ff_cols.index(name) + 1).number_format = "0.000"
            _autofit(ws_ff, len(ff_cols), {1: 14})
            ws_ff.freeze_panes = "A5"
            ws_ff.auto_filter.ref = f"A4:{get_column_letter(len(ff_cols))}{4 + len(ff_factors)}"


        # Strategy Returns: the dependent variable used in the FF5 regression.
        ws_sr = wb.create_sheet("Strategy Returns")
        ws_sr.sheet_view.showGridLines = False
        ws_sr["A1"] = f"{ticker} -- Analyst Sentiment Long-Short Strategy Returns"
        ws_sr["A1"].font = Font(name="Arial", size=14, bold=True)
        ws_sr["A2"] = "Excess return series used as the dependent variable in the Fama-French 5 regression."
        ws_sr["A2"].font = Font(name="Arial", size=9, color="666666")
        sr_cols = [c for c in ["date", "strategy_excess_return"] if c in strategy_returns.columns]
        if sr_cols:
            ws_sr.append([])
            ws_sr.append(sr_cols + ["Cumulative Wealth"])
            _style_header(ws_sr, 3, row=4)
            for i, row in strategy_returns.iterrows():
                excel_row = 5 + i
                date_val = row.get("date", "")
                ret_val = row.get("strategy_excess_return", "")
                ws_sr.cell(excel_row, 1, date_val)
                ws_sr.cell(excel_row, 2, ret_val)
                if excel_row == 5:
                    ws_sr.cell(excel_row, 3, f'=IF(ISNUMBER(B{excel_row}),1+B{excel_row}/100,"")')
                else:
                    ws_sr.cell(excel_row, 3, f'=IF(ISNUMBER(B{excel_row}),C{excel_row-1}*(1+B{excel_row}/100),"")')
                ws_sr.cell(excel_row, 1).number_format = "yyyy-mm-dd"
                ws_sr.cell(excel_row, 2).number_format = "0.000%"
                ws_sr.cell(excel_row, 3).number_format = "0.000x"
            _autofit(ws_sr, 3, {1: 14, 2: 22, 3: 18})
            ws_sr.freeze_panes = "A5"
            ws_sr.auto_filter.ref = f"A4:C{4 + len(strategy_returns)}"


        # FF5 Regression: direct summary plus t-stats where supplied by the
        # regression summary. The table remains auditable because it comes from
        # the same regression object used by the pipeline.
        ws_reg = wb.create_sheet("FF5 Regression")
        ws_reg.sheet_view.showGridLines = False
        ws_reg.merge_cells("A1:D1")
        ws_reg["A1"] = f"{ticker} -- Fama-French 5 Strategy Regression"
        ws_reg["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
        ws_reg["A1"].fill = HEADER_FILL
        ws_reg["A2"] = "Model: pure FF5 = Mkt-RF + SMB + HML + RMW + CMA"
        ws_reg["A2"].font = Font(name="Arial", size=10, color="404040")
        reg_map = {}
        if not factor_regression.empty and {"metric", "value"}.issubset(factor_regression.columns):
            reg_map = dict(zip(factor_regression["metric"], factor_regression["value"]))
        reg_rows = [
            ("Alpha, monthly (%)", reg_map.get("alpha_monthly_pct")),
            ("Alpha t-stat", reg_map.get("alpha_tstat")),
            ("Alpha, annualized (%)", reg_map.get("alpha_annualized_pct")),
            ("Market beta", reg_map.get("loading_Mkt-RF")),
            ("Market beta t-stat", reg_map.get("loading_tstat_Mkt-RF")),
            ("SMB beta", reg_map.get("loading_SMB")),
            ("SMB beta t-stat", reg_map.get("loading_tstat_SMB")),
            ("HML beta", reg_map.get("loading_HML")),
            ("HML beta t-stat", reg_map.get("loading_tstat_HML")),
            ("RMW beta", reg_map.get("loading_RMW")),
            ("RMW beta t-stat", reg_map.get("loading_tstat_RMW")),
            ("CMA beta", reg_map.get("loading_CMA")),
            ("CMA beta t-stat", reg_map.get("loading_tstat_CMA")),
            ("R-squared", reg_map.get("r_squared")),
            ("Observations", reg_map.get("n_obs")),
            ("Mean monthly strategy return (%)", reg_map.get("mean_monthly_return_pct")),
            ("Std monthly strategy return (%)", reg_map.get("std_monthly_return_pct")),
            ("Annualized Sharpe", reg_map.get("sharpe_annualized")),
        ]
        ws_reg.append([])
        ws_reg.append(["Metric", "Value", "Interpretation"])
        _style_header(ws_reg, 3, row=4)
        for idx, (label, value) in enumerate(reg_rows, start=5):
            ws_reg.cell(idx, 1, label)
            ws_reg.cell(idx, 2, value)
            interpretation = ""
            if label == "Alpha, monthly (%)":
                interpretation = "Positive alpha suggests returns not explained by the five systematic factors."
            elif label == "Alpha t-stat":
                interpretation = "Use the t-stat to assess statistical evidence for alpha."
            elif label == "R-squared":
                interpretation = "Share of strategy return variation explained by the FF5 factors."
            elif label == "Observations":
                interpretation = "Monthly observations used in the regression."
            ws_reg.cell(idx, 3, interpretation)
        for r in range(5, 5 + len(reg_rows)):
            if "Alpha" in str(ws_reg.cell(r, 1).value) and "t-stat" not in str(ws_reg.cell(r, 1).value):
                ws_reg.cell(r, 2).number_format = "0.000"
            elif "R-squared" in str(ws_reg.cell(r, 1).value):
                ws_reg.cell(r, 2).number_format = "0.000"
            elif "Observations" in str(ws_reg.cell(r, 1).value):
                ws_reg.cell(r, 2).number_format = "0"
            else:
                ws_reg.cell(r, 2).number_format = "0.000"
        _autofit(ws_reg, 3, {1: 30, 2: 16, 3: 64})
        ws_reg.freeze_panes = "A5"


        # FF5 Diagnostics: data coverage and factor completeness. This stays
        # descriptive and avoids pretending a single regression is sufficient
        # evidence when the strategy history is short.
        ws_diag = wb.create_sheet("FF5 Diagnostics")
        ws_diag.sheet_view.showGridLines = False
        ws_diag.merge_cells("A1:D1")
        ws_diag["A1"] = f"{ticker} -- FF5 Diagnostics"
        ws_diag["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
        ws_diag["A1"].fill = HEADER_FILL
        diagnostics = [
            ("Factor observations", len(ff_factors) if not ff_factors.empty else ""),
            ("Strategy return observations", len(strategy_returns) if not strategy_returns.empty else ""),
            ("Factor start date", ff_factors["date"].min() if "date" in ff_factors.columns and not ff_factors.empty else ""),
            ("Factor end date", ff_factors["date"].max() if "date" in ff_factors.columns and not ff_factors.empty else ""),
            ("Strategy start date", strategy_returns["date"].min() if "date" in strategy_returns.columns and not strategy_returns.empty else ""),
            ("Strategy end date", strategy_returns["date"].max() if "date" in strategy_returns.columns and not strategy_returns.empty else ""),
            ("Regression observations", reg_map.get("n_obs", "")),
            ("Model", "Pure FF5"),
            ("Momentum included", "No"),
        ]
        ws_diag.append([])
        ws_diag.append(["Diagnostic", "Value", "Note"])
        _style_header(ws_diag, 3, row=3)
        for r, (label, value) in enumerate(diagnostics, start=4):
            ws_diag.cell(r, 1, label)
            ws_diag.cell(r, 2, value)
        ws_diag.cell(4, 3, "Counts are taken from the exported factor and strategy panels.")
        ws_diag.cell(7, 3, "Coverage dates are descriptive; regression overlap can be shorter after inner-join/dropna.")
        ws_diag.cell(10, 3, "Pure FF5 excludes MOM. FF5+MOM remains available as a separate model.")
        _autofit(ws_diag, 3, {1: 28, 2: 18, 3: 70})


    out_path = f"{outdir}/live_{safe}_model.xlsx"
    os.makedirs(outdir, exist_ok=True)
    wb.save(out_path)
    return out_path




def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--ticker", default="AAPL-US")
    args = parser.parse_args()
    path = build_workbook(args.ticker)
    print(f"Wrote {path}")
    print("Run scripts/recalc.py against it (see the xlsx skill) if you add/edit formulas later, "
          "or just open it in Excel -- Excel recalculates on open automatically.")




if __name__ == "__main__":
    main()
