"""
Export Smart Consensus outputs to a separate Excel workbook.

Reads:
    outputs/smart_consensus_predictions.csv
    outputs/smart_consensus_analyst_weights.csv
    outputs/smart_consensus_summary.csv

Writes:
    outputs/Smart_Consensus_Model.xlsx

No FactSet/API calls.
"""

from __future__ import annotations

import argparse
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter


OUTPUT_DIR = Path("outputs")
PREDICTIONS = OUTPUT_DIR / "smart_consensus_predictions.csv"
WEIGHTS = OUTPUT_DIR / "smart_consensus_analyst_weights.csv"
SUMMARY = OUTPUT_DIR / "smart_consensus_summary.csv"
ANALYST_AGGREGATE = OUTPUT_DIR / "smart_consensus_analyst_aggregate.csv"
SECTOR_LEADERBOARD = OUTPUT_DIR / "smart_consensus_sector_leaderboard.csv"
BROKER_AGGREGATE = OUTPUT_DIR / "smart_consensus_broker_aggregate.csv"
DEFAULT_XLSX = OUTPUT_DIR / "Smart_Consensus_Model.xlsx"

NAVY = "1F4E78"
BLUE = "5B9BD5"
LIGHT_BLUE = "D9EAF7"
GREEN = "E2F0D9"
LIGHT_GRAY = "E7E6E6"
WHITE = "FFFFFF"
BLACK = "000000"
ORANGE = "FCE4D6"

thin_gray = Side(style="thin", color="BFBFBF")
bottom_border = Border(bottom=thin_gray)


def style_header(ws, row: int, start_col: int = 1, end_col: int | None = None) -> None:
    if end_col is None:
        end_col = ws.max_column
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bottom_border


def autofit(ws, min_width: int = 10, max_width: int = 32) -> None:
    for col_cells in ws.columns:
        col_idx = col_cells[0].column
        max_len = 0
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(max_len + 2, min_width),
            max_width,
        )


def write_dataframe(ws, df: pd.DataFrame, start_row: int = 1) -> None:
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=col_idx, value=col_name)
    style_header(ws, start_row, 1, len(df.columns))

    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws.freeze_panes = f"A{start_row + 1}"
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(df.columns))}{start_row + len(df)}"


def load_csvs() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    missing = [p for p in (PREDICTIONS, WEIGHTS, SUMMARY) if not p.exists()]
    if missing:
        names = ", ".join(str(p) for p in missing)
        raise FileNotFoundError(
            f"Missing Smart Consensus output files: {names}. "
            "Run smart_consensus.py first."
        )

    return (
        pd.read_csv(PREDICTIONS),
        pd.read_csv(WEIGHTS),
        pd.read_csv(SUMMARY),
    )


def load_sector_csvs() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Analyst Aggregate, Sector Leaderboard, and Broker Aggregate are newer,
    optional outputs (added after the original three) -- older
    smart_consensus.py runs won't have written them yet. Missing files
    degrade to empty sheets with a note, rather than crashing the whole
    export.
    """
    analyst_aggregate = pd.read_csv(ANALYST_AGGREGATE) if ANALYST_AGGREGATE.exists() else pd.DataFrame()
    sector_lb = pd.read_csv(SECTOR_LEADERBOARD) if SECTOR_LEADERBOARD.exists() else pd.DataFrame()
    broker_aggregate = pd.read_csv(BROKER_AGGREGATE) if BROKER_AGGREGATE.exists() else pd.DataFrame()
    return analyst_aggregate, sector_lb, broker_aggregate


def build_workbook(output_path: Path) -> Path:
    predictions, weights, summary = load_csvs()
    analyst_aggregate, sector_lb, broker_aggregate = load_sector_csvs()

    # Build a readable analyst contribution summary for each firm-quarter.
    if not weights.empty and {"firm", "quarter", "analyst", "final_weight"}.issubset(weights.columns):
        contrib = weights.copy()
        contrib["final_weight"] = pd.to_numeric(contrib["final_weight"], errors="coerce")
        contrib = contrib.dropna(subset=["final_weight"])
        contrib = contrib.sort_values(
            ["firm", "quarter", "final_weight"],
            ascending=[True, True, False],
        )

        analyst_lists = (
            contrib.groupby(["firm", "quarter"], as_index=False)
            .agg(
                analysts_used=("analyst", lambda s: ", ".join(map(str, s))),
                top_analyst=("analyst", "first"),
                top_weight=("final_weight", "first"),
            )
        )
    else:
        analyst_lists = pd.DataFrame(
            columns=["firm", "quarter", "analysts_used", "top_analyst", "top_weight"]
        )

    if not predictions.empty:
        predictions = predictions.merge(
            analyst_lists,
            on=["firm", "quarter"],
            how="left",
        )
        predictions["analysts_used"] = predictions["analysts_used"].fillna("")
        predictions["top_analyst"] = predictions["top_analyst"].fillna("")
        predictions["top_weight"] = predictions["top_weight"].fillna(0.0)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    # ------------------------------------------------------------
    # Read Me / Methodology
    # ------------------------------------------------------------
    ws = wb.create_sheet("Read Me")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Smart Consensus Model"
    ws["A1"].font = Font(size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)

    ws.merge_cells("A1:F1")

    sections = [
        (
            "What this workbook does",
            "Compares the ordinary Street Consensus with a reliability weighted Smart Consensus using analyst history already collected by the main Analyst Model. It uses existing CSVs only and makes 0 FactSet calls.",
        ),
        (
            "Standard Consensus",
            "The current version uses the median of the available analyst EPS estimates for the company and quarter.",
        ),
        (
            "Smart Consensus",
            "The current version weights analysts using only their historical observations from quarters before the target quarter. Better historical normalized forecast accuracy receives more weight, and analysts with more observations receive more credibility.",
        ),
        (
            "Out of sample rule",
            "An analyst's future performance is never used to set the weight for an earlier target quarter. Only prior quarters are eligible when the target prediction is calculated.",
        ),
        (
            "Accuracy measure",
            "The comparison uses the same normalized forecast error basis as the main model: (forecast EPS minus actual EPS) divided by the stock price 10 trading days before earnings. Lower absolute error is better.",
        ),
        (
            "Credibility",
            "The current credibility adjustment is n observations / (n observations + 10). The 10 is the project's existing credibility prior. Change MIN_HISTORY_OBS in smart_consensus.py to require more history before an analyst can receive a Smart weight.",
        ),
        (
            "Important interpretation",
            "This workbook is a backtest, not a guarantee of future performance. The current dataset is still small, so treat early results as a diagnostic rather than a final conclusion.",
        ),
    ]

    row = 3
    for title, body in sections:
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=1).font = Font(bold=True, color=NAVY)
        ws.cell(row=row + 1, column=1, value=body)
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=6)
        ws.cell(row=row + 1, column=1).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
        row += 3

    ws["A25"] = "How to change minimum analyst history"
    ws["A25"].font = Font(bold=True, color=WHITE)
    ws["A25"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A25:F25")

    ws["A27"] = "Edit this line in smart_consensus.py:"
    ws["A28"] = "MIN_HISTORY_OBS = 4"
    ws["A28"].font = Font(name="Consolas", bold=True, color=BLUE)

    ws["A30"] = "Then rerun:"
    ws["A31"] = "py -3.11 smart_consensus.py --min-history 10"
    ws["A31"].font = Font(name="Consolas", color=BLUE)

    ws["A33"] = "Recommended first comparison"
    ws["A34"] = "Run the same model at 4, 6, 8, and 10 prior observations. Do not select a threshold simply because it gives the best result on a tiny in-sample sample. The point is to see whether Smart Consensus becomes more stable as the evidence requirement increases."
    ws["A34"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A34:F35")

    ws.column_dimensions["A"].width = 28
    for c in range(2, 7):
        ws.column_dimensions[get_column_letter(c)].width = 18
    for r in range(1, 36):
        ws.row_dimensions[r].height = 20

    # ------------------------------------------------------------
    # Dashboard
    # ------------------------------------------------------------
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Smart Consensus Dashboard"
    ws["A1"].font = Font(size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:F1")

    if not summary.empty:
        for r_idx, row_data in enumerate(summary.itertuples(index=False), start=3):
            metric = row_data[0]
            standard = row_data[1]
            smart = row_data[2]
            ws.cell(r_idx, 1, metric)
            ws.cell(r_idx, 2, standard)
            ws.cell(r_idx, 3, smart)

        style_header(ws, 2, 1, 3)
        ws["A2"] = "Metric"
        ws["B2"] = "Standard"
        ws["C2"] = "Smart"

        for r in range(3, ws.max_row + 1):
            metric = str(ws.cell(r, 1).value)
            if "error" in metric.lower():
                ws.cell(r, 2).number_format = "0.0000"
                ws.cell(r, 3).number_format = "0.0000"
            elif "rate" in metric.lower():
                ws.cell(r, 2).number_format = "0.0%"
                ws.cell(r, 3).number_format = "0.0%"
            else:
                ws.cell(r, 2).number_format = "0.00"
                ws.cell(r, 3).number_format = "0.00"

    # Coverage block
    ws["E2"] = "Coverage"
    ws["E2"].fill = PatternFill("solid", fgColor=NAVY)
    ws["E2"].font = Font(color=WHITE, bold=True)
    if not predictions.empty and "smart_available" in predictions.columns:
        total = len(predictions)
        available = int(predictions["smart_available"].sum())
        ws["E3"] = "All target quarters"
        ws["F3"] = total
        ws["E4"] = "Smart available"
        ws["F4"] = available
        ws["E5"] = "Smart coverage"
        ws["F5"] = (available / total) if total else 0
        ws["F5"].number_format = "0.0%"
    else:
        ws["E3"] = "No prediction data"

    ws["E7"] = "Interpretation"
    ws["E7"].fill = PatternFill("solid", fgColor=NAVY)
    ws["E7"].font = Font(color=WHITE, bold=True)
    ws["E2"].fill = PatternFill("solid", fgColor=NAVY)
    ws["E2"].font = Font(color=WHITE, bold=True)
    ws["E8"] = "Lower MAE is better."
    ws["E9"] = "Higher win rate is better."
    ws["E10"] = "Treat the result as a backtest, not a guarantee."
    for r in range(8, 11):
        ws.cell(r, 5).alignment = Alignment(wrap_text=True)

    if not summary.empty and {"metric", "standard", "smart"}.issubset(summary.columns):
        mae = summary[summary["metric"] == "Mean absolute forecast error"]
        if not mae.empty:
            chart = BarChart()
            chart.title = "Mean Absolute Forecast Error"
            chart.y_axis.title = "Normalized absolute FE"
            chart.x_axis.title = "Method"
            data = Reference(ws, min_col=2, max_col=3, min_row=2, max_row=3)
            cats = Reference(ws, min_col=1, min_row=3, max_row=3)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 7
            chart.width = 12
            ws.add_chart(chart, "E8")

    autofit(ws)

    # ------------------------------------------------------------
    # Predictions
    # ------------------------------------------------------------
    ws = wb.create_sheet("Predictions")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Standard vs Smart Consensus"
    ws["A1"].font = Font(size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(8, len(predictions.columns)))

    ws["A2"] = (
        "Use Firm + Quarter to trace a Smart Consensus directly to the named analysts below. "
        "If Smart is unavailable, the model did not have enough historical analyst evidence."
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(8, len(predictions.columns)))
    ws["A2"].alignment = Alignment(wrap_text=True)

    write_dataframe(ws, predictions, start_row=4)

    # Number formatting
    for row in range(5, ws.max_row + 1):
        for col_name in [
            "standard_consensus",
            "smart_consensus",
            "actual_eps",
        ]:
            if col_name in predictions.columns:
                col = list(predictions.columns).index(col_name) + 1
                ws.cell(row, col).number_format = "0.0000"

        for col_name in [
            "standard_fe",
            "smart_fe",
            "standard_abs_fe",
            "smart_abs_fe",
            "smart_weight_coverage",
        ]:
            if col_name in predictions.columns:
                col = list(predictions.columns).index(col_name) + 1
                ws.cell(row, col).number_format = "0.0000"

    if "top_weight" in predictions.columns:
        col = list(predictions.columns).index("top_weight") + 1
        for row in range(5, ws.max_row + 1):
            ws.cell(row, col).number_format = "0.0%"

    # Highlight winners
    if "winner" in predictions.columns:
        winner_col = list(predictions.columns).index("winner") + 1
        for row in range(5, ws.max_row + 1):
            cell = ws.cell(row, winner_col)
            if cell.value == "Smart":
                cell.fill = PatternFill("solid", fgColor=GREEN)
            elif cell.value == "Standard":
                cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    # ------------------------------------------------------------
    # Analyst Weights
    # ------------------------------------------------------------
    ws = wb.create_sheet("Analyst Weights")
    ws.sheet_view.showGridLines = False
    write_dataframe(ws, weights)

    if not weights.empty:
        for col_name in [
            "historical_mae",
            "accuracy_component",
            "credibility_weight",
            "normalized_weight",
            "final_weight",
        ]:
            if col_name in weights.columns:
                col = list(weights.columns).index(col_name) + 1
                for row in range(2, ws.max_row + 1):
                    ws.cell(row, col).number_format = "0.0000"

    # ------------------------------------------------------------
    # Analyst Aggregate -- every analyst's FULL track record pooled
    # across every ticker she covers in this pull, not per-firm. This is
    # a plain in-sample summary (mean absolute forecast error over her
    # whole history) -- easier to scan than the out-of-sample Analyst
    # Weights sheet above, which only shows the weight she actually
    # earned for each specific prediction.
    # ------------------------------------------------------------
    ws = wb.create_sheet("Analyst Aggregate")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Analyst Aggregate -- full track record across every ticker covered"
    ws["A1"].font = Font(size=16, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(8, len(analyst_aggregate.columns) if not analyst_aggregate.empty else 8))
    ws["A2"] = (
        "avg_abs_fe: lower is better (mean absolute forecast error across ALL her observations, every ticker). "
        "n_times_smart_weighted / avg_weight_when_smart_weighted: how much real influence she earned in the "
        "out-of-sample Smart Consensus blend above -- 0 means she never had enough prior history to be used."
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(8, len(analyst_aggregate.columns) if not analyst_aggregate.empty else 8))
    ws["A2"].alignment = Alignment(wrap_text=True)
    if analyst_aggregate.empty:
        ws["A4"] = "No data -- run smart_consensus.py (this repo's newer version) to produce smart_consensus_analyst_aggregate.csv."
        ws["A4"].font = Font(italic=True, color="C00000")
    else:
        write_dataframe(ws, analyst_aggregate, start_row=4)
        if "avg_abs_fe" in analyst_aggregate.columns:
            col = list(analyst_aggregate.columns).index("avg_abs_fe") + 1
            for row in range(5, ws.max_row + 1):
                ws.cell(row, col).number_format = "0.0000"
        if "avg_weight_when_smart_weighted" in analyst_aggregate.columns:
            col = list(analyst_aggregate.columns).index("avg_weight_when_smart_weighted") + 1
            for row in range(5, ws.max_row + 1):
                ws.cell(row, col).number_format = "0.0%"
    autofit(ws)

    # ------------------------------------------------------------
    # Sector Leaderboard -- analysts ranked WITHIN their industry
    # (Fama-French 48 group, resolved from SIC code) instead of across
    # the whole universe. n_tickers_in_sector tells you honestly whether
    # a given sector's ranking means anything yet (1 ticker = not a real
    # sector comparison, it's just that ticker's own leaderboard).
    # ------------------------------------------------------------
    ws = wb.create_sheet("Sector Leaderboard")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Sector Leaderboard -- best analysts WITHIN each industry"
    ws["A1"].font = Font(size=16, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(8, len(sector_lb.columns) if not sector_lb.empty else 8))
    ws["A2"] = (
        "Check n_tickers_in_sector before trusting a ranking as a real sector comparison: a sector with only 1 "
        "ticker pulled so far is just that ticker's own analysts wearing a sector label, not a true peer comparison. "
        "As you pull more tickers per industry, these become genuinely comparative."
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(8, len(sector_lb.columns) if not sector_lb.empty else 8))
    ws["A2"].alignment = Alignment(wrap_text=True)
    if sector_lb.empty:
        ws["A4"] = "No data -- run smart_consensus.py (this repo's newer version) to produce smart_consensus_sector_leaderboard.csv."
        ws["A4"].font = Font(italic=True, color="C00000")
    else:
        write_dataframe(ws, sector_lb, start_row=4)
        if "avg_abs_fe" in sector_lb.columns:
            col = list(sector_lb.columns).index("avg_abs_fe") + 1
            for row in range(5, ws.max_row + 1):
                ws.cell(row, col).number_format = "0.0000"
        # Flag thin (1-ticker) sectors so nobody mistakes them for a real comparison.
        if "n_tickers_in_sector" in sector_lb.columns:
            col = list(sector_lb.columns).index("n_tickers_in_sector") + 1
            for row in range(5, ws.max_row + 1):
                cell = ws.cell(row, col)
                if cell.value == 1:
                    cell.fill = PatternFill("solid", fgColor=ORANGE)
    autofit(ws)

    # ------------------------------------------------------------
    # Broker Aggregate -- rank BROKERAGE FIRMS, not individuals, pooling
    # every analyst at that house across every ticker (not sector-specific).
    # Answers "which house's research is actually most accurate" as a firm-
    # level question, e.g. is KeyBanc's semiconductor/industrials coverage
    # genuinely more reliable across the board than BMO's.
    # ------------------------------------------------------------
    ws = wb.create_sheet("Broker Aggregate")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Broker Aggregate -- how good is each brokerage firm, full stop"
    ws["A1"].font = Font(size=16, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(8, len(broker_aggregate.columns) if not broker_aggregate.empty else 8))
    ws["A2"] = (
        "avg_abs_fe: lower is better, pooled across every analyst at that broker and every ticker they cover -- "
        "not sector-specific. Check n_analysts before trusting a broker's rank: n_analysts=1 is really just one "
        "person's track record wearing the firm's name, while 5+ is a genuine house-wide signal."
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(8, len(broker_aggregate.columns) if not broker_aggregate.empty else 8))
    ws["A2"].alignment = Alignment(wrap_text=True)
    if broker_aggregate.empty:
        ws["A4"] = "No data -- run smart_consensus.py (this repo's newer version) to produce smart_consensus_broker_aggregate.csv."
        ws["A4"].font = Font(italic=True, color="C00000")
    else:
        write_dataframe(ws, broker_aggregate, start_row=4)
        if "avg_abs_fe" in broker_aggregate.columns:
            col = list(broker_aggregate.columns).index("avg_abs_fe") + 1
            for row in range(5, ws.max_row + 1):
                ws.cell(row, col).number_format = "0.0000"
        # Flag single-analyst brokers so nobody mistakes one person's track
        # record for a firm-wide result.
        if "n_analysts" in broker_aggregate.columns:
            col = list(broker_aggregate.columns).index("n_analysts") + 1
            for row in range(5, ws.max_row + 1):
                cell = ws.cell(row, col)
                if cell.value == 1:
                    cell.fill = PatternFill("solid", fgColor=ORANGE)
    autofit(ws)

    # ------------------------------------------------------------
    # Consensus Detail
    # ------------------------------------------------------------
    ws = wb.create_sheet("Consensus Detail")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "How Smart Consensus Is Built"
    ws["A1"].font = Font(size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:J1")

    ws["A3"] = (
        "Each row below is one analyst contributing to a Smart Consensus. "
        "Filter Firm + Quarter to see exactly whose estimates were used, "
        "their historical evidence, and their final weight."
    )
    ws.merge_cells("A3:J4")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")

    detail_cols = [
        c
        for c in [
            "firm",
            "quarter",
            "analyst",
            "broker",
            "broker_code",
            "estimate_value",
            "n_prior_obs",
            "historical_mae",
            "credibility_weight",
            "final_weight",
        ]
        if c in weights.columns
    ]

    detail_df = weights[detail_cols].copy() if detail_cols else pd.DataFrame()

    if not detail_df.empty:
        write_dataframe(ws, detail_df, start_row=6)

        for col_name in [
            "estimate_value",
            "historical_mae",
            "credibility_weight",
            "final_weight",
        ]:
            if col_name in detail_df.columns:
                col = list(detail_df.columns).index(col_name) + 1
                for r in range(7, 7 + len(detail_df)):
                    ws.cell(r, col).number_format = "0.0000"

        if "credibility_weight" in detail_df.columns:
            col = list(detail_df.columns).index("credibility_weight") + 1
            for r in range(7, 7 + len(detail_df)):
                ws.cell(r, col).number_format = "0.0%"

        if "final_weight" in detail_df.columns:
            col = list(detail_df.columns).index("final_weight") + 1
            for r in range(7, 7 + len(detail_df)):
                ws.cell(r, col).number_format = "0.0%"
    else:
        ws["A6"] = "No Smart Consensus analyst weights are currently available."
        ws["A6"].fill = PatternFill("solid", fgColor=ORANGE)

    autofit(ws)

    # ------------------------------------------------------------
    # Diagnostics
    # ------------------------------------------------------------
    ws = wb.create_sheet("Diagnostics")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Smart Consensus Diagnostics"
    ws["A1"].font = Font(size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:H1")

    # Overall coverage diagnostics
    total_targets = len(predictions)
    smart_available = (
        int(predictions["smart_available"].sum())
        if not predictions.empty and "smart_available" in predictions.columns
        else 0
    )
    smart_unavailable = total_targets - smart_available
    smart_coverage = smart_available / total_targets if total_targets else np.nan

    diagnostics = [
        ("All target firm quarters", total_targets),
        ("Smart Consensus available", smart_available),
        ("Smart Consensus unavailable", smart_unavailable),
        ("Smart coverage", smart_coverage),
        (
            "Unique firms",
            predictions["firm"].nunique() if not predictions.empty else 0,
        ),
        (
            "Unique quarters",
            predictions["quarter"].nunique() if not predictions.empty else 0,
        ),
        (
            "Unique analysts receiving Smart weights",
            weights["analyst"].nunique()
            if not weights.empty and "analyst" in weights.columns
            else 0,
        ),
        (
            "Average current analysts",
            predictions["n_current_estimates"].mean()
            if not predictions.empty and "n_current_estimates" in predictions.columns
            else np.nan,
        ),
        (
            "Average Smart weighted analysts",
            predictions["n_smart_weighted_analysts"].mean()
            if not predictions.empty and "n_smart_weighted_analysts" in predictions.columns
            else np.nan,
        ),
        (
            "Average Smart weight coverage when available",
            predictions.loc[
                predictions["smart_available"], "smart_weight_coverage"
            ].mean()
            if not predictions.empty and "smart_available" in predictions.columns
            else np.nan,
        ),
    ]

    ws["A3"] = "Overall Diagnostic"
    ws["B3"] = "Value"
    style_header(ws, 3, 1, 2)

    for r, (label, value) in enumerate(diagnostics, start=4):
        ws.cell(r, 1, label)
        ws.cell(r, 2, value)

        if "coverage" in label.lower():
            ws.cell(r, 2).number_format = "0.0%"
        elif isinstance(value, float):
            ws.cell(r, 2).number_format = "0.00"

    # Eligibility / unavailable reason at the target level
    ws["D3"] = "Coverage by Target"
    ws["D3"].font = Font(color=WHITE, bold=True)
    ws["D3"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("D3:H3")

    coverage_cols = [
        c
        for c in [
            "firm",
            "quarter",
            "n_current_estimates",
            "n_smart_weighted_analysts",
            "smart_weight_coverage",
            "smart_available",
            "winner",
        ]
        if c in predictions.columns
    ]

    if coverage_cols:
        coverage_view = predictions[coverage_cols].copy()

        # Add a plain English status.
        coverage_view["status"] = np.where(
            coverage_view["smart_available"],
            "Smart available",
            "Not enough analyst history",
        )

        # Put status first for readability.
        ordered = ["firm", "quarter", "status"] + [
            c for c in coverage_view.columns
            if c not in {"firm", "quarter", "status"}
        ]
        coverage_view = coverage_view[ordered]

        write_dataframe(
            ws,
            coverage_view,
            start_row=4,
        )

        # The write_dataframe helper will overwrite the existing header area
        # in row 4 onward, so adjust the title after writing.
        ws["D3"] = "Coverage by Target"
        ws["D3"].font = Font(color=WHITE, bold=True)
        ws["D3"].fill = PatternFill("solid", fgColor=NAVY)
        ws.merge_cells(
            start_row=3,
            start_column=4,
            end_row=3,
            end_column=min(4 + len(coverage_view.columns) - 1, 8),
        )

        # Format coverage column if present.
        if "smart_weight_coverage" in coverage_view.columns:
            c = list(coverage_view.columns).index("smart_weight_coverage") + 1
            for rr in range(5, 5 + len(coverage_view)):
                ws.cell(rr, 4 - 1 + c).number_format = "0.0%"

    # Why Smart is unavailable
    ws["A17"] = "Why Smart Can Be Unavailable"
    ws["A17"].font = Font(color=WHITE, bold=True)
    ws["A17"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A17:H17")

    notes = [
        "Smart Consensus requires at least MIN_HISTORY_OBS prior forecast-error observations for an analyst.",
        "If no current analyst has enough prior history, Standard Consensus is still shown but Smart Consensus is blank.",
        "This is a coverage limitation, not a failed Standard Consensus calculation.",
        "As the master company universe grows and analysts accumulate more cross-company history, Smart coverage should increase.",
        "Do not choose a minimum-history threshold simply because it gives the lowest error on a tiny sample. Compare thresholds out of sample on a larger dataset.",
    ]

    for i, note in enumerate(notes, start=19):
        ws.cell(i, 1, f"• {note}")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
        ws.cell(i, 1).alignment = Alignment(wrap_text=True)

    # Target-level table of Smart unavailable rows
    if not predictions.empty and "smart_available" in predictions.columns:
        unavailable = predictions[
            ~predictions["smart_available"]
        ].copy()

        if not unavailable.empty:
            ws["A27"] = "Targets Without Smart Consensus"
            ws["A27"].font = Font(color=WHITE, bold=True)
            ws["A27"].fill = PatternFill("solid", fgColor=ORANGE)
            ws.merge_cells("A27:H27")

            unavailable_cols = [
                c
                for c in [
                    "firm",
                    "quarter",
                    "n_current_estimates",
                    "n_smart_weighted_analysts",
                    "smart_weight_coverage",
                    "winner",
                ]
                if c in unavailable.columns
            ]

            unavailable_view = unavailable[unavailable_cols].copy()
            unavailable_view["reason"] = "No analyst met minimum history requirement"
            unavailable_view = unavailable_view[
                ["firm", "quarter", "reason"]
                + [
                    c
                    for c in unavailable_view.columns
                    if c not in {"firm", "quarter", "reason"}
                ]
            ]

            write_dataframe(
                ws,
                unavailable_view,
                start_row=28,
            )

    # ------------------------------------------------------------
    # Threshold Guide
    # ------------------------------------------------------------
    ws = wb.create_sheet("Threshold Guide")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Minimum History Threshold Guide"
    ws["A1"].font = Font(size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:F1")

    ws["A3"] = "Current code setting"
    ws["A3"].font = Font(bold=True)
    ws["B3"] = "MIN_HISTORY_OBS = 4"
    ws["B3"].font = Font(name="Consolas", color=BLUE)

    ws["A5"] = "Run these separately"
    ws["A5"].font = Font(color=WHITE, bold=True)
    ws["A5"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A5:F5")

    thresholds = [4, 6, 8, 10]

    for idx, threshold in enumerate(thresholds, start=7):
        ws.cell(idx, 1, f"Threshold {threshold}")
        ws.cell(
            idx,
            2,
            f"py -3.11 smart_consensus.py --min-history {threshold}",
        )
        ws.cell(idx, 2).font = Font(name="Consolas", color=BLUE)

    ws["A13"] = "How to interpret the test"
    ws["A13"].font = Font(color=WHITE, bold=True)
    ws["A13"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A13:F13")

    ws["A15"] = (
        "Higher thresholds demand stronger analyst history, which can reduce "
        "coverage while potentially making Smart weights more reliable. "
        "Do not pick a threshold only because it wins on the current tiny "
        "sample. The useful question is whether Smart remains better as the "
        "sample grows and the threshold is fixed before evaluation."
    )
    ws.merge_cells("A15:F17")
    ws["A15"].alignment = Alignment(wrap_text=True, vertical="top")

    autofit(ws)


    # Global workbook settings
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = ws.freeze_panes or "A2"

    # Patrick only wants Analyst Aggregate and Sector Leaderboard visible day
    # to day -- those are the two views built for "rank analysts per sector
    # and per accuracy, and map analysts covering multiple companies".
    # Everything else here (Read Me, Dashboard, Predictions, Analyst Weights,
    # Consensus Detail, Diagnostics, Threshold Guide) is hidden rather than
    # deleted -- it's the out-of-sample Standard-vs-Smart comparison and its
    # supporting methodology/notes, which only covers tickers with the
    # newer raw_estimates/raw_actuals/raw_prices files (fewer tickers than
    # Analyst Aggregate/Sector Leaderboard, which is why it can look
    # confusingly limited if you're on the wrong tab). Unhide any of these
    # from Excel any time: right-click a visible sheet tab -> Unhide.
    for _hidden_sheet in (
        "Read Me", "Dashboard", "Predictions", "Analyst Weights",
        "Consensus Detail", "Diagnostics", "Threshold Guide",
    ):
        if _hidden_sheet in wb.sheetnames:
            wb[_hidden_sheet].sheet_state = "hidden"
    # Excel requires at least one visible sheet with an active tab; make
    # sure the workbook opens straight to Analyst Aggregate.
    if "Analyst Aggregate" in wb.sheetnames:
        wb.active = wb.sheetnames.index("Analyst Aggregate")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export Smart Consensus outputs to Excel. No FactSet calls."
    )
    parser.add_argument(
        "--output",
        default=str(DEFAULT_XLSX),
        help="Output xlsx path.",
    )
    args = parser.parse_args()

    path = build_workbook(Path(args.output))
    print("FactSet/API calls: 0")
    print(f"Wrote {path}")


if __name__ == "__main__":
    main()
